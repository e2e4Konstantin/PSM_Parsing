
from openpyxl.worksheet import worksheet
from openpyxl.utils.cell import column_index_from_string
from openpyxl.styles import DEFAULT_FONT
import openpyxl

from db.get_data_db import get_data_db, get_db_data_with_header
from files_tolls import output_message_exit, file_unused
from db.sql_queries import select_query



def _basic_header_output(sheet: worksheet):
    sheet.append(["."])
    header = ["Шифр ПСМ", "Наименование ПСМ", "Измеритель ПСМ ",
              "Номер позиции в ПСМ", "Шифр позиции", "Наименование ресурса", "Измеритель",
              'сопоставление параметров и атрибутов', 'формула объема', 'пояснения/шифр расценки',
              'критерии/сценарии', 'проверка на пилоте',
              'файл', 'row_number']
    sheet.append(header)
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']:
        sheet.column_dimensions[col].width = 25


def _data_line_output(sheet: worksheet, src_data: tuple, start_row: int) -> int:
    sheet.cell(row=start_row, column=column_index_from_string('A')).value = src_data[0]
    sheet.cell(row=start_row, column=column_index_from_string('B')).value = src_data[1]
    sheet.cell(row=start_row, column=column_index_from_string('C')).value = src_data[2]
    sheet.cell(row=start_row, column=column_index_from_string('D')).value = src_data[3]
    sheet.cell(row=start_row, column=column_index_from_string('E')).value = src_data[4]
    sheet.cell(row=start_row, column=column_index_from_string('F')).value = src_data[5]
    sheet.cell(row=start_row, column=column_index_from_string('G')).value = src_data[6]
    sheet.cell(row=start_row, column=column_index_from_string('H')).value = src_data[7]
    sheet.cell(row=start_row, column=column_index_from_string('I')).value = src_data[8]
    sheet.cell(row=start_row, column=column_index_from_string('J')).value = src_data[9]
    sheet.cell(row=start_row, column=column_index_from_string('K')).value = src_data[10]
    sheet.cell(row=start_row, column=column_index_from_string('L')).value = src_data[11]
    sheet.cell(row=start_row, column=column_index_from_string('M')).value = src_data[12]
    sheet.cell(row=start_row, column=column_index_from_string('N')).value = src_data[13]
    return 1


def _statistics_line_output(sheet: worksheet, src_data: tuple, start_row: int) -> int:
    sheet.cell(row=start_row, column=column_index_from_string('A')).value = src_data[0]
    sheet.cell(row=start_row, column=column_index_from_string('B')).value = src_data[1]
    sheet.cell(row=start_row, column=column_index_from_string('C')).value = src_data[2]
    sheet.cell(row=start_row, column=column_index_from_string('D')).value = src_data[3]
    sheet.cell(row=start_row, column=column_index_from_string('E')).value = src_data[4]
    sheet.cell(row=start_row, column=column_index_from_string('F')).value = src_data[5]
    sheet.cell(row=start_row, column=column_index_from_string('G')).value = src_data[6]
    sheet.cell(row=start_row, column=column_index_from_string('H')).value = src_data[7]

    return 1


def write_excel_file(excel_file_name: str, path_db: str):
    if file_unused(excel_file_name):
        book, sheet = None, None
        try:
            book = openpyxl.Workbook()
            DEFAULT_FONT.font = "Calibri"
            DEFAULT_FONT.sz = 8
            sheet = book.active # получить ссылку на активную книгу
            sheet.title = "data P6"
            book.create_sheet("all")
            book.create_sheet("stat")

        except IOError as err:
            output_message_exit(f"Ошибка при создании excel файла: {excel_file_name!r}", f"{err}")
        # получить данные из БД для Р6
        data_set = get_data_db(path_db, select_query["select_P6_raw_parsers"])
        if data_set:
            # вывод шапки таблицы
            _basic_header_output(sheet)
            x = 4
            for row in data_set:
                x += _data_line_output(sheet, row[1:], x)

        # получить ВСЕ данные из БД6
        data_set = get_data_db(path_db, select_query["select_all_raw_parsers"])
        if data_set:
            sheet_all = book["all"]
            _basic_header_output(sheet_all)
            x = 4
            for row in data_set:
                x += _data_line_output(sheet_all, row[1:], x)

        # получить статистику
        data_set = get_db_data_with_header(path_db, select_query["select_count_P_raw_parsers"])
        if data_set:
            sheet_all = book["stat"]
            x = 4
            for row in data_set:
                x += _statistics_line_output(sheet_all, row, x)


        if book:
            book.save(excel_file_name)
            book.close()
    else:
        output_message_exit(f"Файл занят другим приложением: ", f"{excel_file_name!r}")






if __name__ == "__main__":
    db_path = r"F:\Kazak\GoogleDrive\1_KK\Job_CNAC\Python_projects\development\PSM_Parsing\output\PSM.sqlite3"
    write_excel_file("test.xlsx", db_path)