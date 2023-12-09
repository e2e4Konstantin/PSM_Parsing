import os
import re

from icecream import ic
from openpyxl.worksheet import worksheet
from openpyxl import load_workbook, Workbook
from openpyxl.utils.cell import column_index_from_string
from files_tolls import output_message_exit, output_message

from config import DataLine


def _get_data_line(work_sheet: worksheet, target_row: int) -> tuple:
    psm_code = work_sheet['D2'].value
    psm_description = work_sheet['D3'].value
    psm_measure = work_sheet['D4'].value

    item_number = work_sheet.cell(row=target_row, column=column_index_from_string('A')).value
    item_code = work_sheet.cell(row=target_row, column=column_index_from_string('B')).value
    item_description = work_sheet.cell(row=target_row, column=column_index_from_string('C')).value
    item_measure = work_sheet.cell(row=target_row, column=column_index_from_string('D')).value
    comparison = work_sheet.cell(row=target_row, column=column_index_from_string('E')).value
    volume_formula = work_sheet.cell(row=target_row, column=column_index_from_string('F')).value
    notes = work_sheet.cell(row=target_row, column=column_index_from_string('G')).value

    criteria = work_sheet.cell(row=target_row, column=column_index_from_string('M')).value
    inspection = work_sheet.cell(row=target_row, column=column_index_from_string('N')).value

    return (psm_code, psm_description, psm_measure,
            item_number, item_code, item_description, item_measure,
            comparison, volume_formula, notes, criteria, inspection)


#
# def get_data_from_excel(file_name: str) -> list[tuple]:
#     book, sheet = None, None
#     sheet_name = "ОГ"
#     mask_name = re.compile(r"^\s*(ОГ)|(OГ)\s*")
#     marker = "Р6"
#     result = []
#     try:
#         book = load_workbook(filename=file_name, data_only=True)
#         sheets = book.sheetnames
#         ts = [x for x in sheets if mask_name.match(x)]
#         if len(ts) > 0:
#             book[ts[0]].title = sheet_name
#             sheet = book[sheet_name]
#             # ic(sheet.title)
#             # ic(sheet.max_row, sheet.max_column)
#             column_m = column_index_from_string('M')
#             target_rows = [row
#                            for row in range(1, sheet.max_row + 1)
#                            if sheet.cell(row=row, column=column_m).value == marker
#                            ]
#             base_file_name = os.path.basename(file_name)
#             ic(base_file_name, target_rows)
#             for row in target_rows:
#                 result.append(_get_data_line(sheet, row)+(base_file_name, ))
#             # ic(result)
#         else:
#             output_message(f"В excel файле: {file_name!r}", f"нет страницы: {sheet_name!r}")
#         book.close()
#         return result
#     except IOError as err:
#         output_message_exit(f"Ошибка при открытии excel файла: {file_name!r}", f"{err}")
#     return []
#

def get_focus_sheet(wbook: Workbook) -> worksheet:
    """ Ищет в книге wbook страницу с имнем похожим на 'ОГ'. """
    if wbook:
        sheet_name = "ОГ"
        mask_name = re.compile(r"^\s*(ОГ)|(OГ)\s*")
        focus_name = [x for x in wbook.sheetnames if mask_name.match(x)]
        if len(focus_name) > 0:
            wbook[focus_name[0]].title = sheet_name
            return wbook[sheet_name]
    return None


def detect_data_limits(wsheet: worksheet) -> tuple[int, int] | None:
    """ Определяет границы таблицы с данными.
        Ищет шапку для начала таблицы и пустую строку для окончания.
    """
    a_header = "№"
    b_header = "Шифр таблицы/группы/расценки или ресурса"
    c_header = "Наименование"
    d_header = "Измеритель"
    end_columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'M', 'N']
    start_row, end_row = 0, 0
    row_max = wsheet.max_row
    for row in range(1, row_max + 1):
        header_condition = [
            a_header == wsheet.cell(row=row, column=column_index_from_string('A')).value,
            b_header == wsheet.cell(row=row, column=column_index_from_string('B')).value,
            c_header == wsheet.cell(row=row, column=column_index_from_string('C')).value,
            d_header == wsheet.cell(row=row, column=column_index_from_string('D')).value,
        ]
        if all(header_condition):
            start_row = row
        end_value = [wsheet.cell(row=row, column=column_index_from_string(column)).value for column in end_columns]

        end_condition = [(x == "" or x is None) for x in end_value]
        if (start_row > 0 and all(end_condition)) or row == row_max:
            if row == row_max:
                end_row = row
            else:
                end_row = row - 1
            break
    if start_row > 0 and end_row > start_row:
        return start_row, end_row
    return None


def get_data_items_from_excel(file_name: str) -> list[tuple]:
    result = []
    try:
        book = load_workbook(filename=file_name, data_only=True)
        sheet = get_focus_sheet(book)
        if sheet:
            limits = detect_data_limits(sheet)
            if limits:
                base_file_name = os.path.basename(file_name)
                for row in range(limits[0] + 1, limits[1] + 1):
                    result.append(_get_data_line(sheet, row) + (base_file_name, row,))
            else:
                output_message(f"В excel файле: {file_name!r} на страницы: {sheet.title!r}", "нет таблицы с данными")
        else:
            output_message(f"В excel файле: {file_name!r}", f"нет страницы: {book.sheetnames}")
        book.close()
        return result
    except IOError as err:
        output_message_exit(f"Ошибка при открытии excel файла: {file_name!r}", f"{err}")
    return []
