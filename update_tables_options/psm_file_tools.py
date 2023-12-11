from icecream import ic
from openpyxl.worksheet import worksheet
from openpyxl import load_workbook, Workbook
from openpyxl.utils.cell import column_index_from_string
from openpyxl.styles import Color, Font

import re

from files_tolls import output_message_exit
from update_tables_options.db_attributes import get_db_attributes_for_table
from update_tables_options.db_options import get_db_options_for_table
from config import dbTolls


def get_focus_sheet(wbook: Workbook) -> worksheet:
    """ Ищет в книге wbook страницу с имнем похожим на 'ОГ'. """
    if wbook:
        sheet_name = "ОГ"
        mask_name = re.compile(r"^\s*(ОГ)|(OГ)\s*")
        focus_name = [x for x in wbook.sheetnames if mask_name.match(x)]
        if len(focus_name) > 0:
            wbook[focus_name[0]].title = sheet_name
            return wbook[sheet_name]
    return None


def get_bulletin_tables(src_sheet: worksheet) -> tuple[tuple[int, str, str], ...]:
    if src_sheet:
        row_max = src_sheet.max_row
        # ic(sheet.title, row_max)
        tables = []
        table_marker = re.compile(r"^\d+$")
        table_code = re.compile(r"^\d+\.\d+-\d+$")

        for row in range(1, row_max + 1):
            value_a = str(src_sheet.cell(row=row, column=column_index_from_string('A')).value).strip()
            value_b = str(src_sheet.cell(row=row, column=column_index_from_string('B')).value).strip()
            if table_marker.match(value_a) and table_code.match(value_b):
                tables.append((row, value_a, value_b))
        return tuple(tables)
    return tuple()


def psm_restore(psm_file_name: str, db_name: str):
    try:
        book = load_workbook(filename=psm_file_name, data_only=True)
        sheet = get_focus_sheet(book)
        if sheet:
            tables = get_bulletin_tables(src_sheet=sheet)
            ic(psm_file_name, tables)
            with dbTolls(db_name) as db:
                for table in tables:
                    message = []
                    ic(table)
                    attributes = get_db_attributes_for_table(db, table[2])
                    options = get_db_options_for_table(db, table[2])

                    message.extend(sorted(attributes.keys())) if attributes is not None else None
                    message.extend(sorted(options.keys())) if options is not None else None
                    # ic(message)
                    if not message:
                        break
                    row = table[0]
                    column_number = column_index_from_string('H')
                    changed_string = ', '.join(message)
                    font = Font(name='Calibri', color=Color(rgb='00CD5C5C'), size=11, bold=True)

                    sheet.cell(row=row, column=column_number).value = changed_string
                    sheet.cell(row=row, column=column_number).font = font
                    ic(changed_string)
                book.save(psm_file_name)
        else:
            output_message_exit(f"В excel файле: {psm_file_name!r}", f"нет страницы 'ОГ'")
        book.close()
    except IOError as err:
        output_message_exit(f"Ошибка при открытии excel файла: {psm_file_name!r}", f"{err}")
